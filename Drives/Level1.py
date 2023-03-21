import openpyxl as xl
import os
class Searchdrives():
    """
    This module helps to find all drives present in pc
    67 to 93 for ascii value chr()
    to convert ascii value to char
    """
    def __init__(self):
        self.drives=[]
        self.workbook=xl.load_workbook("c://testdata//testdrive.xlsx")
    def searchmydrives(self):
        for x in range(67,91):
            if os.path.exists(chr(x)+":"):
                self.drives.append(chr(x))
        return self.drives
dr=Searchdrives()
data=str(dr.searchmydrives())
print(data)
sheet=dr.workbook.active
sheet.cell(row=1,column=1).value=data
sheet.cell(row=2,column=1).value=data
sheet.cell(row=3,column=1).value=data
dr.workbook.save("c://testdata//testdrive.xlsx")
dr.workbook.close()